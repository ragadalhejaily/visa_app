from flask import Flask, render_template, request, redirect, session
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os

app = Flask(__name__)
app.secret_key = "secret_key_for_session"
excel_file = 'database.xlsx'

def init_excel():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "رحلات"
        ws.append(["تاريخ الذهاب", "تاريخ العودة"])
        wb.save(excel_file)

def read_trips():
    wb = load_workbook(excel_file)
    ws = wb.active
    trips = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            try:
                start = row[0]
                end = row[1]
                if not isinstance(start, datetime):
                    start = datetime.strptime(str(start), "%Y-%m-%d")
                if not isinstance(end, datetime):
                    end = datetime.strptime(str(end), "%Y-%m-%d")
                trips.append((start.date(), end.date()))
            except Exception as e:
                print(f"خطأ في قراءة الصف: {row}, الخطأ: {e}")
    return trips

def add_trip(start_date, end_date):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.append([start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")])
    wb.save(excel_file)

def delete_trip(index):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.delete_rows(index + 2)
    wb.save(excel_file)

def calculate_used_days(trips, window_start, window_end):
    total_days = 0
    for start, end in trips:
        if end < window_start or start > window_end:
            continue
        actual_start = max(start, window_start)
        actual_end = min(end, window_end)
        duration = (actual_end - actual_start).days
        # duration = (actual_end - actual_start).days + 1
        total_days += duration
    return total_days

@app.route("/", methods=["GET", "POST"])
def index():
    init_excel()
    trips = read_trips()

    if request.method == "POST":
        # حذف رحلة
        if "delete_index" in request.form:
            delete_index = int(request.form["delete_index"])
            delete_trip(delete_index)
            return redirect("/")

        # إضافة رحلة جديدة
        elif "new_trip_start" in request.form and "new_trip_end" in request.form:
            start = datetime.strptime(request.form["new_trip_start"], "%Y-%m-%d").date()
            end = datetime.strptime(request.form["new_trip_end"], "%Y-%m-%d").date()

            if start > end:
                session['result'] = {"warning": "⚠️ تاريخ الذهاب يجب أن يكون قبل تاريخ العودة"}
            else:
                window_start = end - timedelta(days=180)
                used_days = calculate_used_days(trips, window_start, end)
                # new_trip_days = (end - start).days + 1
                new_trip_days = (end - start).days
                total_used = used_days + new_trip_days
                remaining = 90 - total_used
                warning = "✅ نظامي" if remaining >= 0 else f"⚠️ غير نظامي  تجاوزت الحد بـ {abs(remaining)} يوم"

                # إضافة الرحلة الجديدة إلى ملف الإكسل بعد التحقق (يمكن تعديل حسب رغبتك)
                add_trip(start, end)

                session['result'] = {
                    "window_start": window_start.strftime('%Y-%m-%d'),
                    "used_days": used_days,
                    "new_trip_days": new_trip_days,
                    "total_used": total_used,
                    "remaining": remaining,
                    "warning": warning,
                    "check_start": start.strftime('%Y-%m-%d'),
                    "check_end": end.strftime('%Y-%m-%d')
                }
            return redirect("/")

        # تحقق من رحلة جديدة (دون الإضافة)
        elif "check_start" in request.form and "check_end" in request.form:
            start = datetime.strptime(request.form["check_start"], "%Y-%m-%d").date()
            end = datetime.strptime(request.form["check_end"], "%Y-%m-%d").date()

            if start > end:
                session['result'] = {"warning": "⚠️ تاريخ الذهاب يجب أن يكون قبل تاريخ العودة"}
            else:
                window_start = end - timedelta(days=180)
                used_days = calculate_used_days(trips, window_start, end)
                # new_trip_days = (end - start).days + 1
                new_trip_days = (end - start).days
                total_used = used_days + new_trip_days
                remaining = 90 - total_used
                warning = "✅ نظامي" if remaining >= 0 else f"⚠️غير نظامي تجاوزت الحد بـ {abs(remaining)} يوم"

                session['result'] = {
                    "window_start": window_start.strftime('%Y-%m-%d'),
                    "used_days": used_days,
                    "new_trip_days": new_trip_days,
                    "total_used": total_used,
                    "remaining": remaining,
                    "warning": warning,
                    "check_start": start.strftime('%Y-%m-%d'),
                    "check_end": end.strftime('%Y-%m-%d')
                }
            return redirect("/")

    result = session.pop("result", None)
    return render_template("index.html", trips=trips, result=result)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)